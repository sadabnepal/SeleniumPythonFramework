import openpyxl


class ExcelReader:

    @staticmethod
    def get_excel_data_as_dictionary(test_name):
        workbook = openpyxl.load_workbook("..//TestData//test_data.xlsx")
        sheet = workbook.active
        Dict = {}
        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_name:
                for j in range(2, sheet.max_row+1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
